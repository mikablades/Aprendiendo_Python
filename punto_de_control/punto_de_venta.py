"""_summary_
    Formulario.
"""
import tkinter as tk
from tkinter import ttk
import openpyxl
from ttkthemes import ThemedStyle


class FormularioApp:
    """_summary_
    1. lee o crea un .xlsx
    2. agrega encabezados
    3. agrega datos
    4. guarda datos sin reemplazar los preexistentes
    5. anula la ultima entrada
    """

    def __init__(self, root):
        self.lista_datos = root
        self.boton_anular = root
        self.boton_guardar = root
        self.combo_articulo = root
        self.etiqueta_articulo = root
        self.entry_cantidad = root
        self.etiqueta_cantidad = root
        self.entry_valor_compra = root
        self.etiqueta_valor_compra = root
        self.entry_valor_venta = root
        self.etiqueta_valor_venta = root
        self.menu_ayuda = root
        self.menu_archivo = root
        self.barra_menus = root
        self.marco = root
        self.root = root
        self.etiqueta_descripcion = root
        self.entry_descripcion = root
        
        root.title("Punto de Venta Sereno.cl")

        # Intentar cargar un archivo Excel existente o crear uno nuevo si no existe
        try:
            self.archivo_excel = openpyxl.load_workbook("inv.xlsx")
        except FileNotFoundError:
            self.archivo_excel = openpyxl.Workbook()

        # Utilizar la primera hoja del archivo Excel
        self.hoja = self.archivo_excel.active

        # Si la hoja está vacía, agregar encabezados
        if self.hoja.max_row == 1:
            self.hoja.append(["Articulo", "Descripcion", "Cantidad", "Valor compra", "Valor Venta", "Margen", "Estado"])

        self.create_gui()

    def create_gui(self):
        """_summary_
        Interfaz gráfica (aún no me decido por el estilo, ¿hay más estilos para tk?)
        """
        # marco
        self.marco = ttk.Frame(self.root)
        self.marco.pack(padx=1, pady=1, fill="both", expand=True)

        style = ThemedStyle(self.root)
        style.set_theme("black")

        self.marco["style"] = "Black.TFrame"
        # Menu
        self.barra_menus = tk.Menu(self.root)
        self.root.config(menu=self.barra_menus)

        self.menu_archivo = tk.Menu(self.barra_menus, tearoff=False)
        self.barra_menus.add_cascade(menu=self.menu_archivo, label="Archivo")

        self.menu_ayuda = tk.Menu(self.barra_menus, tearoff=False)
        self.barra_menus.add_cascade(menu=self.menu_ayuda, label="Ayuda")

        self.menu_archivo.add_command(label="Nuevo", accelerator="Ctrl+N", command=self.root)
        self.menu_ayuda.add_command(label="ayudaaa!", accelerator="Ctrl+Y", command=self.root)
        
        #Articulo        
        self.etiqueta_articulo = ttk.Label(self.marco, text="Articulo")
        self.etiqueta_articulo.grid(row=0, column=0, padx=5, pady=5, sticky="w")

        self.combo_articulo = ttk.Combobox(self.marco,
                                         state="readonly",
                                         values=["MagKiss", "Magnet Stick 26", "Magnet Blocks 40", "DouDaner", "Platos y cubiertos", "Fijadores de Ceja", "Vasos Lugano"])
        self.combo_articulo.grid(row=0, column=1, padx=5, pady=5)
        
        # descripcion
        self.etiqueta_descripcion = ttk.Label(self.marco, text="Descripcion:")
        self.etiqueta_descripcion.grid(row=1, column=0, padx=5, pady=5, sticky="w")

        self.entry_descripcion = ttk.Entry(self.marco)
        self.entry_descripcion.grid(row=1, column=1, padx=5, pady=5)
        # valor_compra
        self.etiqueta_valor_compra = ttk.Label(self.marco, text="Valor Compra")
        self.etiqueta_valor_compra.grid(row=2, column=0, padx=5, pady=5, sticky="w")

        self.entry_valor_compra = ttk.Entry(self.marco)
        self.entry_valor_compra.grid(row=2, column=1, padx=5, pady=5)
        
        # valor_venta
        self.etiqueta_valor_venta = ttk.Label(self.marco, text="Valor venta:")
        self.etiqueta_valor_venta.grid(row=3, column=0, padx=5, pady=5, sticky="w")

        self.entry_valor_venta = ttk.Entry(self.marco)
        self.entry_valor_venta.grid(row=3, column=1, padx=5, pady=5)
        # cantidad
        self.etiqueta_cantidad = ttk.Label(self.marco, text="Cantidad:")
        self.etiqueta_cantidad.grid(row=4, column=0, padx=5, pady=5, sticky="w")

        self.entry_cantidad = ttk.Entry(self.marco)
        self.entry_cantidad.grid(row=4, column=1, padx=5, pady=5)
    
        

        # lista de datos en pantalla
        columnas = ("Articulo", "Descripcion", "Cantidad", "Valor compra", "Valor Venta", "Margen", "Estado")
        self.lista_datos = ttk.Treeview(self.marco, columns=columnas, show="headings")
        for columna in columnas:
            self.lista_datos.heading(columna, text=columna)
            self.lista_datos.column(columna, width=100)
        self.lista_datos.grid(row=5, column=0, columnspan=2, padx=5, pady=5)

        for fila in self.hoja.iter_rows(min_row=2, values_only=True):
            self.lista_datos.insert("", "end", values=fila)
        # boton "guardar"
        self.boton_guardar = ttk.Button(self.marco, text="Guardar", command=self.guardar_datos)
        self.boton_guardar.grid(row=6, column=0, padx=5, pady=10)
        style.configure("TButton", borderwidth=0, padding=5, relief="flat")
        # boton "anular"
        self.boton_anular = ttk.Button(self.marco, text="Anular última entrada",
                                       command=self.anular_ultima_entrada)
        self.boton_anular.grid(row=6, column=1, padx=5, pady=10)

    def guardar_datos(self):
        """_summary_
        Guarda los datos
        """
        articulo = self.combo_articulo.get()
        descripcion = self.entry_descripcion.get()
        cantidad = self.entry_cantidad.get()
        valor_compra = self.entry_valor_compra.get()
        valor_venta = self.entry_valor_venta.get()
        

        fila = [articulo, descripcion, cantidad, valor_compra, valor_venta]
        self.hoja.append(fila)
        self.archivo_excel.save("inv.xlsx")

        self.lista_datos.insert("", "end", values=(articulo, descripcion, cantidad, valor_compra, valor_venta,))
        self.entry_valor_venta.delete(0, "end")
        self.entry_cantidad.delete(0, "end")
        self.entry_valor_compra.delete(0, "end")
        self.combo_articulo.delete(0, "end")
        self.entry_descripcion.delete(0, "end")

    def anular_ultima_entrada(self):
        """_summary_
        borra el ultimo dato de cada columna
        """
        if self.hoja.max_row > 1:
            self.hoja.delete_rows(self.hoja.max_row)
            self.lista_datos.delete(*self.lista_datos.get_children())


if __name__ == "__main__":
    ventana_principal = tk.Tk()
    app = FormularioApp(ventana_principal)
    ventana_principal.mainloop()

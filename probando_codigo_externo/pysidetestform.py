import sys
import openpyxl
from PySide6.QtWidgets import QApplication, QMainWindow, QLabel, QLineEdit, QPushButton, QVBoxLayout, QWidget, QTreeView, QStandardItemModel, QStandardItem

class FormularioApp:
    def __init__(self, root):
        self.root = root
        root.setWindowTitle("Formulario")

        # Intentar cargar un archivo Excel existente o crear uno nuevo si no existe
        try:
            self.archivo_excel = openpyxl.load_workbook("datos.xlsx")
        except FileNotFoundError:
            self.archivo_excel = openpyxl.Workbook()

        # Utilizar la primera hoja del archivo Excel
        self.hoja = self.archivo_excel.active

        # Si la hoja está vacía, agregar encabezados
        if self.hoja.max_row == 1:
            self.hoja.append(["Nombre", "Edad"])

        self.create_gui()

    def create_gui(self):
        self.marco = QWidget(self.root)

        self.layout = QVBoxLayout(self.marco)
        self.root.setCentralWidget(self.marco)

        self.etiqueta_nombre = QLabel("Nombre:")
        self.layout.addWidget(self.etiqueta_nombre)

        self.entry_nombre = QLineEdit()
        self.layout.addWidget(self.entry_nombre)

        self.etiqueta_edad = QLabel("Edad:")
        self.layout.addWidget(self.etiqueta_edad)

        self.entry_edad = QLineEdit()
        self.layout.addWidget(self.entry_edad)

        self.boton_guardar = QPushButton("Guardar")
        self.layout.addWidget(self.boton_guardar)
        self.boton_guardar.clicked.connect(self.guardar_datos)

        self.boton_anular = QPushButton("Anular última entrada")
        self.layout.addWidget(self.boton_anular)
        self.boton_anular.clicked.connect(self.anular_ultima_entrada)

        columnas = ["Nombre", "Edad"]
        self.modelo = QStandardItemModel()
        self.modelo.setHorizontalHeaderLabels(columnas)

        self.lista_datos = QTreeView()
        self.lista_datos.setModel(self.modelo)
        self.layout.addWidget(self.lista_datos)

        for fila in self.hoja.iter_rows(min_row=2, values_only=True):
            self.insertar_fila_en_lista(fila)

    def guardar_datos(self):
        nombre = self.entry_nombre.text()
        edad = self.entry_edad.text()

        fila = [nombre, edad]
        self.hoja.append(fila)
        self.archivo_excel.save("datos.xlsx")

        self.insertar_fila_en_lista(fila)
        self.entry_nombre.clear()
        self.entry_edad.clear()

    def anular_ultima_entrada(self):
        if self.hoja.max_row > 1:
            self.hoja.delete_rows(self.hoja.max_row)
            self.modelo.clear()
            self.modelo.setHorizontalHeaderLabels(["Nombre", "Edad"])

    def insertar_fila_en_lista(self, fila):
        items = [QStandardItem(str(dato)) for dato in fila]
        self.modelo.appendRow(items)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    ventana_principal = QMainWindow()
    formulario = FormularioApp(ventana_principal)
    ventana_principal.show()
    sys.exit(app.exec_())
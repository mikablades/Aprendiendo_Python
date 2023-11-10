"""_summary_
    Formulario.
"""
import tkinter as tk
from tkinter import ttk
import openpyxl
from ttkthemes import ThemedStyle


class FormularioApp:
    """_summary_
    1. lee o crea un .xlsx
    2. agrega encabezados
    3. agrega datos
    4. guarda datos sin reemplazar los preexistentes
    5. anula la ultima entrada
    """

    def __init__(self, root):
        self.lista_datos = root
        self.boton_anular = root
        self.boton_guardar = root
        self.combo_genero = root
        self.etiqueta_genero = root
        self.entry_edad = root
        self.etiqueta_edad = root
        self.entry_rut = root
        self.etiqueta_rut = root
        self.entry_nombre = root
        self.etiqueta_nombre = root
        self.menu_ayuda = root
        self.menu_archivo = root
        self.barra_menus = root
        self.marco = root
        self.root = root
        root.title("Formulario")

        # Intentar cargar un archivo Excel existente o crear uno nuevo si no existe
        try:
            self.archivo_excel = openpyxl.load_workbook("datos.xlsx")
        except FileNotFoundError:
            self.archivo_excel = openpyxl.Workbook()

        # Utilizar la primera hoja del archivo Excel
        self.hoja = self.archivo_excel.active

        # Si la hoja está vacía, agregar encabezados
        if self.hoja.max_row == 1:
            self.hoja.append(["Nombre", "Edad", "Rut", "Genero"])

        self.create_gui()

    def create_gui(self):
        """_summary_
        Interfaz gráfica (aún no me decido por el estilo, ¿hay más estilos para tk?)
        """
        # marco
        self.marco = ttk.Frame(self.root)
        self.marco.pack(padx=1, pady=1, fill="both", expand=True)

        style = ThemedStyle(self.root)
        style.set_theme("black")

        self.marco["style"] = "Black.TFrame"
        # Menu
        self.barra_menus = tk.Menu(self.root)
        self.root.config(menu=self.barra_menus)

        self.menu_archivo = tk.Menu(self.barra_menus, tearoff=False)
        self.barra_menus.add_cascade(menu=self.menu_archivo, label="Archivo")

        self.menu_ayuda = tk.Menu(self.barra_menus, tearoff=False)
        self.barra_menus.add_cascade(menu=self.menu_ayuda, label="Ayuda")

        self.menu_archivo.add_command(label="Nuevo", accelerator="Ctrl+N", command=self.root)
        # nombre
        self.etiqueta_nombre = ttk.Label(self.marco, text="Nombre:")
        self.etiqueta_nombre.grid(row=0, column=0, padx=5, pady=5, sticky="w")

        self.entry_nombre = ttk.Entry(self.marco)
        self.entry_nombre.grid(row=0, column=1, padx=5, pady=5)
        # rut
        self.etiqueta_rut = ttk.Label(self.marco, text="Rut:")
        self.etiqueta_rut.grid(row=1, column=0, padx=5, pady=5, sticky="w")

        self.entry_rut = ttk.Entry(self.marco)
        self.entry_rut.grid(row=1, column=1, padx=5, pady=5)
        # edad
        self.etiqueta_edad = ttk.Label(self.marco, text="Edad:")
        self.etiqueta_edad.grid(row=2, column=0, padx=5, pady=5, sticky="w")

        self.entry_edad = ttk.Entry(self.marco)
        self.entry_edad.grid(row=2, column=1, padx=5, pady=5)
        # Genero
        self.etiqueta_genero = ttk.Label(self.marco, text="Genero:")
        self.etiqueta_genero.grid(row=3, column=0, padx=5, pady=5, sticky="w")

        self.combo_genero = ttk.Combobox(self.marco,
                                         state="readonly",
                                         values=["hombre", "mujer", "no binario", "genero fluido"])
        self.combo_genero.grid(row=3, column=1, padx=5, pady=5)

        # boton "guardar"
        self.boton_guardar = ttk.Button(self.marco, text="Guardar", command=self.guardar_datos)
        self.boton_guardar.grid(row=5, column=0, padx=5, pady=10)
        style.configure("TButton", borderwidth=0, padding=5, relief="flat")
        # boton "anular"
        self.boton_anular = ttk.Button(self.marco, text="Anular última entrada",
                                       command=self.anular_ultima_entrada)
        self.boton_anular.grid(row=5, column=1, padx=5, pady=10)
        # lista de datos en pantalla
        columnas = ("Nombre", "Edad", "Rut", "Genero")
        self.lista_datos = ttk.Treeview(self.marco, columns=columnas, show="headings")
        for columna in columnas:
            self.lista_datos.heading(columna, text=columna)
            self.lista_datos.column(columna, width=100)
        self.lista_datos.grid(row=4, column=0, columnspan=2, padx=5, pady=5)

        for fila in self.hoja.iter_rows(min_row=2, values_only=True):
            self.lista_datos.insert("", "end", values=fila)

    def guardar_datos(self):
        """_summary_
        Guarda los datos
        """
        nombre = self.entry_nombre.get()
        edad = self.entry_edad.get()
        rut = self.entry_rut.get()
        genero = self.combo_genero.get()

        fila = [nombre, edad, rut, genero]
        self.hoja.append(fila)
        self.archivo_excel.save("datos.xlsx")

        self.lista_datos.insert("", "end", values=(nombre, edad, rut, genero))
        self.entry_nombre.delete(0, "end")
        self.entry_edad.delete(0, "end")
        self.entry_rut.delete(0, "end")
        self.combo_genero.delete(0, "end")

    def anular_ultima_entrada(self):
        """_summary_
        borra el ultimo dato de cada columna
        """
        if self.hoja.max_row > 1:
            self.hoja.delete_rows(self.hoja.max_row)
            self.lista_datos.delete(*self.lista_datos.get_children())


if __name__ == "__main__":
    ventana_principal = tk.Tk()
    app = FormularioApp(ventana_principal)
    ventana_principal.mainloop()

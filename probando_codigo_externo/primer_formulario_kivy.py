"""_summary_

    Returns:
        _type_: _description_
"""    
import pandas as pd
from kivy.app import App
from kivy.uix.spinner import Spinner
from kivy.uix.button import Button
from kivy.uix.boxlayout import BoxLayout
import openpyxl

def __init__(self, root):
    self.root = root
    root.title("Formulario")

        # Intentar cargar un archivo Excel existente o crear uno nuevo si no existe
    try:
        self.archivo_excel = openpyxl.load_workbook("datos.xlsx")
    except FileNotFoundError:
        self.archivo_excel = openpyxl.Workbook()

        # Utilizar la primera hoja del archivo Excel
    self.hoja = self.archivo_excel.active

        # Si la hoja está vacía, agregar encabezados
    if self.hoja.max_row == 1:
        self.hoja.append(["Variable 1", "Variable 2", "Variable 3"])

    self.build(self)

class Formulario(App):
    """_summary_

    Args:
        App (_type_): _description_
    """
def build(self):
    """_summary_

    Returns:
        _type_: _description_
    """
    self.df = pd.DataFrame(columns=['Variable 1', 'Variable 2', 'Variable 3'])
    self.layout = BoxLayout(orientation='vertical')
    self.spinner1 = Spinner(text='Variable 1', values=('Opción 1', 'Opción 2', 'Opción 3'))
    self.spinner2 = Spinner(text='Variable 2', values=('Opción 1', 'Opción 2', 'Opción 3'))
    self.spinner3 = Spinner(text='Variable 3', values=('Opción 1', 'Opción 2', 'Opción 3'))
    self.button = Button(text='Enviar', on_press=self.guardar_datos)
    self.layout.add_widget(self.spinner1)
    self.layout.add_widget(self.spinner2)
    self.layout.add_widget(self.spinner3)
    self.layout.add_widget(self.button)
    return self.layout

def guardar_datos(self, instance):
    datos = {'Variable 1': [self.spinner1.text], 
             'Variable 2': [self.spinner2.text], 
             'Variable 3': [self.spinner3.text]}
    df_temp = pd.DataFrame(datos)
    self.df = self.df.append(df_temp, ignore_index=True)
    self.df.to_excel('datos.xlsx', index=False)

if __name__ == '__main__':
    Formulario().run()
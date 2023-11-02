import tkinter as tk
from tkinter import ttk
import openpyxl
from ttkthemes import ThemedStyle

class FormularioApp:
    def __init__(self, root):
        self.root = root
        root.title("Formulario")

        # Intentar cargar un archivo Excel existente o crear uno nuevo si no existe
        try:
            self.archivo_excel = openpyxl.load_workbook("datos.xlsx")
        except FileNotFoundError:
            self.archivo_excel = openpyxl.Workbook()

        # Utilizar la primera hoja del archivo Excel
        self.hoja = self.archivo_excel.active

        # Si la hoja está vacía, agregar encabezados
        if self.hoja.max_row == 1:
            self.hoja.append(["Nombre", "Edad"])

        self.create_gui()

    def create_gui(self):
        self.marco = ttk.Frame(self.root)
        self.marco.pack(padx=10, pady=10, fill="both", expand=True)

        style = ThemedStyle(self.root)
        style.set_theme("scidgreen")

        self.marco["style"] = "Scidgreen.TFrame"

        self.etiqueta_nombre = ttk.Label(self.marco, text="Nombre:")
        self.etiqueta_nombre.grid(row=0, column=0, padx=5, pady=5, sticky="w")

        self.entry_nombre = ttk.Entry(self.marco)
        self.entry_nombre.grid(row=0, column=1, padx=5, pady=5)

        self.etiqueta_edad = ttk.Label(self.marco, text="Edad:")
        self.etiqueta_edad.grid(row=1, column=0, padx=5, pady=5, sticky="w")

        self.entry_edad = ttk.Entry(self.marco)
        self.entry_edad.grid(row=1, column=1, padx=5, pady=5)

        self.boton_guardar = ttk.Button(self.marco, text="Guardar", command=self.guardar_datos)
        self.boton_guardar.grid(row=2, column=0, padx=5, pady=10)
        style.configure("TButton", borderwidth=0, padding=5, relief="flat")

        self.boton_anular = ttk.Button(self.marco, text="Anular última entrada", command=self.anular_ultima_entrada)
        self.boton_anular.grid(row=2, column=1, padx=5, pady=10)

        columnas = ("Nombre", "Edad")
        self.lista_datos = ttk.Treeview(self.marco, columns=columnas, show="headings")
        for columna in columnas:
            self.lista_datos.heading(columna, text=columna)
            self.lista_datos.column(columna, width=100)
        self.lista_datos.grid(row=3, column=0, columnspan=2, padx=5, pady=5)

        for fila in self.hoja.iter_rows(min_row=2, values_only=True):
            self.lista_datos.insert("", "end", values=fila)

    def guardar_datos(self):
        nombre = self.entry_nombre.get()
        edad = self.entry_edad.get()

        fila = [nombre, edad]
        self.hoja.append(fila)
        self.archivo_excel.save("datos.xlsx")

        self.lista_datos.insert("", "end", values=(nombre, edad))
        self.entry_nombre.delete(0, "end")
        self.entry_edad.delete(0, "end")

    def anular_ultima_entrada(self):
        if self.hoja.max_row > 1:
            self.hoja.delete_rows(self.hoja.max_row)
            self.lista_datos.delete(*self.lista_datos.get_children())

if __name__ == "__main__":
    root = tk.Tk()
    app = FormularioApp(root)
    root.mainloop()


